from __future__ import annotations

import argparse
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from mrna_half_life_paper.config import METADATA_DIR, PROCESSED_DIR, RAW_DIR, ROOT
from mrna_half_life_paper.saluki_naming import saluki_label_column, saluki_label_filename


REAL_DATA_URLS = {
    "article": "https://genomebiology.biomedcentral.com/articles/10.1186/s13059-022-02811-x",
    "github_repo_api": "https://api.github.com/repos/vagarwal87/saluki_paper",
    "zenodo_record_api": "https://zenodo.org/api/records/7158836",
    "moesm2": (
        "https://static-content.springer.com/esm/art%3A10.1186%2Fs13059-022-02811-x/"
        "MediaObjects/13059_2022_2811_MOESM2_ESM.xlsx"
    ),
    "moesm3": (
        "https://static-content.springer.com/esm/art%3A10.1186%2Fs13059-022-02811-x/"
        "MediaObjects/13059_2022_2811_MOESM3_ESM.xlsx"
    ),
}

SUPPLEMENT_CONFIG = {
    "moesm2": {
        "path": RAW_DIR / "supplements" / "13059_2022_2811_MOESM2_ESM.xlsx",
        "header_row": 1,
        "pc1_column": None,
        "matrix_filename": "moesm2_raw_sparse_matrix.tsv",
    },
    "moesm3": {
        "path": RAW_DIR / "supplements" / "13059_2022_2811_MOESM3_ESM.xlsx",
        "header_row": 3,
        "pc1_column": "half-life (PC1)",
        "matrix_filename": "moesm3_saluki_processed_matrix.tsv",
    },
}


def _portable_path(path: Path) -> str:
    """Prefer repository-relative paths in generated provenance tables."""
    try:
        return str(path.resolve().relative_to(ROOT.resolve()))
    except ValueError:
        return str(path)


def parse_sample_id(sample_id: str, species: str) -> dict[str, str]:
    parts = sample_id.split("_")
    if len(parts) < 3:
        raise ValueError(f"Unexpected sample id format: {sample_id}")

    replicate = "_".join(parts[3:]) if len(parts) > 3 else "1"
    return {
        "sample_id": sample_id,
        "study": parts[0],
        "method": parts[1],
        "cell_type": parts[2],
        "replicate": replicate,
        "species": species,
    }


def _load_sheet_dataframe(xlsx_path: Path, sheet_name: str, header_row: int) -> pd.DataFrame:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(min_row=header_row, values_only=True)
    raw_header = next(rows)
    header = [str(value).strip() if value is not None else f"unnamed_{idx}" for idx, value in enumerate(raw_header)]
    df = pd.DataFrame.from_records(list(rows), columns=header)
    df = df.loc[df["Ensembl Gene Id"].notna()].copy()
    return df


def load_supplement(
    species: str,
    supplement_key: str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.Series | None]:
    if supplement_key not in SUPPLEMENT_CONFIG:
        raise ValueError(f"Unknown supplement key: {supplement_key}")

    config = SUPPLEMENT_CONFIG[supplement_key]
    df = _load_sheet_dataframe(config["path"], species, config["header_row"])
    df = df.rename(columns={"Ensembl Gene Id": "gene_id", "Gene name": "gene_name"})

    fixed_cols = {"gene_id", "gene_name"}
    pc1_column = config["pc1_column"]
    if pc1_column is not None:
        fixed_cols.add(pc1_column)

    sample_cols = [col for col in df.columns if col not in fixed_cols]
    matrix = df.loc[:, ["gene_id", *sample_cols]].copy()
    matrix = matrix.mask(matrix.eq("NA"))
    for column in sample_cols:
        matrix[column] = pd.to_numeric(matrix[column], errors="coerce")

    genes = df.loc[:, ["gene_id", "gene_name"]].copy()
    metadata = pd.DataFrame([parse_sample_id(sample, species) for sample in sample_cols])
    saluki_label = None
    if pc1_column is not None:
        saluki_label = pd.to_numeric(df[pc1_column], errors="coerce")
        saluki_label.index = df["gene_id"].astype(str)
        saluki_label = saluki_label.rename(saluki_label_column(species))

    return matrix, genes, metadata, saluki_label


def extract_real_data(
    output_root: Path | None = None,
    metadata_root: Path | None = None,
) -> pd.DataFrame:
    output_root = output_root if output_root is not None else PROCESSED_DIR / "real_data"
    metadata_root = metadata_root if metadata_root is not None else METADATA_DIR
    output_root.mkdir(parents=True, exist_ok=True)
    metadata_root.mkdir(parents=True, exist_ok=True)

    manifest_rows: list[dict[str, str | int | float]] = []

    for species in ("human", "mouse"):
        species_dir = output_root / species
        species_dir.mkdir(parents=True, exist_ok=True)

        species_metadata_written = False
        for supplement_key, config in SUPPLEMENT_CONFIG.items():
            matrix, genes, metadata, saluki_label = load_supplement(species, supplement_key)

            matrix_path = species_dir / config["matrix_filename"]
            gene_path = species_dir / f"{supplement_key}_gene_annotations.tsv"
            matrix.to_csv(matrix_path, sep="\t", index=False)
            genes.to_csv(gene_path, sep="\t", index=False)

            if not species_metadata_written:
                metadata_path = metadata_root / f"real_{species}_samples.tsv"
                metadata.to_csv(metadata_path, sep="\t", index=False)
                species_metadata_written = True

            saluki_label_path = None
            if saluki_label is not None:
                saluki_label_path = species_dir / saluki_label_filename(species)
                saluki_label.to_frame().reset_index().rename(columns={"index": "gene_id"}).to_csv(
                    saluki_label_path,
                    sep="\t",
                    index=False,
                )

            manifest_rows.append(
                {
                    "species": species,
                    "dataset": supplement_key,
                    "n_genes": int(matrix.shape[0]),
                    "n_samples": int(matrix.shape[1] - 1),
                    "matrix_path": _portable_path(matrix_path),
                    "gene_annotation_path": _portable_path(gene_path),
                    "metadata_path": _portable_path(metadata_root / f"real_{species}_samples.tsv"),
                    "saluki_label_path": "" if saluki_label_path is None else _portable_path(saluki_label_path),
                    "source_xlsx": _portable_path(config["path"]),
                }
            )

    manifest = pd.DataFrame(manifest_rows)
    manifest.to_csv(output_root / "manifest.tsv", sep="\t", index=False)

    provenance = {
        "sources": REAL_DATA_URLS,
        "notes": {
            "moesm2": "Public sparse matrix of transformed half-life values collected from published studies.",
            "moesm3": "Public Saluki filtered/transformed/imputed/quantile-normalized matrix with human PC1 consensus.",
        },
    }
    with open(output_root / "provenance.json", "w", encoding="utf-8") as handle:
        json.dump(provenance, handle, indent=2)

    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract the public Saluki half-life supplements into TSV files.")
    parser.add_argument(
        "--output-root",
        type=Path,
        default=PROCESSED_DIR / "real_data",
        help="Directory where extracted real-data tables will be written.",
    )
    parser.add_argument(
        "--metadata-root",
        type=Path,
        default=METADATA_DIR,
        help="Directory where sample metadata tables will be written.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    manifest = extract_real_data(args.output_root, args.metadata_root)
    print(args.output_root)
    print(manifest.to_string(index=False))


if __name__ == "__main__":
    main()
